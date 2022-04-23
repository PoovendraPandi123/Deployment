import openpyxl
import shutil

excel_columns = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "R", "S", "T", "U", "V", "W", "X", "Y", "Z"]

def template_fill(excel_file, sheet_name, cell_name, text):
    try:
        workbook = openpyxl.load_workbook(excel_file)
        worksheet = workbook[sheet_name]

        worksheet[cell_name] = text

        # worksheet_chg = worksheet[cell_name]
        # worksheet_chg.font = Font(color='00FF0000', bold=True)

        workbook.save(excel_file)
        workbook.close()
    except Exception as e:
        return {"Status": "Error", "Message": str(e)}

def template_cell_fill(page):
    try:
        workbook = openpyxl.load_workbook(page["excel_file"])
        worksheet = workbook[page["sheet_name"]]

        for data in page["data"]:
            # print(data["cell_number"])
            # print(data["cell_value"])
            worksheet[data["cell_number"]] = data["cell_value"]

        workbook.save(page["excel_file"])
        workbook.close()
    except Exception as e:
        print("Error")
        print(e)
        return {"Status": "Error", "Message": str(e)}

def template_multiple_fill(excel_file, sheet_name, cell_name, cell_start_number, data_list):
    try:
        workbook = openpyxl.load_workbook(excel_file)
        worksheet = workbook[sheet_name]

        # cell_position = cell_name + str(cell_start_number)

        for i in range(0, len(data_list)):
            # print(cell_name + str(cell_start_number))
            worksheet[cell_name + str(cell_start_number)] = data_list[i]
            cell_start_number = cell_start_number + 1

        workbook.save(excel_file)
        workbook.close()
    except Exception as e:
        print("Error")
        print(e)
        print(data_list[i])
        print(cell_start_number)
        print(cell_name)
        print({"Status": "Error", "Message": str(e)})
        return {"Status": "Error", "Message": str(e)}

def write_vrs_file(data):
    try:
        excel_file = "G:/AdventsProduct/V1.1.0/AFS/VendorReconciliation/static/VRS-Report-Template.xlsx"

        new_file = "G:/AdventsProduct/V1.1.0/AFS/VendorReconciliation/static/" + "VRS-Report-" + str(data["vendor_code"]) + "-" + data["report_generation_count"] + ".xlsx"

        shutil.copy(excel_file, new_file)

        # For Home Page
        home_page = {
            "excel_file": new_file,
            "sheet_name": "Home",
            "data": [
                {"cell_number": "D6", "cell_value": data["report_generation_date"]},
                {"cell_number": "D9", "cell_value": data["vendor_name"]},
                {"cell_number": "D10", "cell_value": data["report_from_date"]},
                {"cell_number": "D11", "cell_value": data["report_to_date"]},
                {"cell_number": "D12", "cell_value": data["vendor_code"]},
                {"cell_number": "D13", "cell_value": data["vendor_site_code"]},
                {"cell_number": "D14", "cell_value": data["vendor_category"]},
                {"cell_number": "D15", "cell_value": data["liability_account"]},
                {"cell_number": "D16", "cell_value": data["division"]},
                {"cell_number": "D17", "cell_value": data["pan_number"]},
                {"cell_number": "D18", "cell_value": data["gst_number"]},
            ]
        }

        template_cell_fill(home_page)

        # For Thermax DR & CR
        for i in range(0, len(data["vrs_rep_tmx_dr_cr_query_output"].columns)):
            template_multiple_fill(new_file, "Thermax DR & CR", excel_columns[i], 7, data["vrs_rep_tmx_dr_cr_query_output"][i])

        # For Vendor DR & CR
        for i in range(0, len(data["vrs_rep_vendor_dr_cr_query_output"].columns)):
            template_multiple_fill(new_file, "Vendor DR & CR", excel_columns[i], 7, data["vrs_rep_vendor_dr_cr_query_output"][i])

        # For Vendor Statement
        for i in range(0, len(data["vrs_rep_vendor_all_query_output"].columns)):
            template_multiple_fill(new_file, "Vendor Statement", excel_columns[i], 4, data["vrs_rep_vendor_all_query_output"][i])

        # For Thermax Statement
        for i in range(0, len(data["vrs_rep_tmx_all_query_output"].columns)):
            template_multiple_fill(new_file, "Thermax Statement", excel_columns[i], 2, data["vrs_rep_tmx_all_query_output"][i])

        file_generated = "http://localhost:50013/static/" + new_file.split("/")[-1]

        return {"Status": "Success", "file_generated": file_generated}
    except Exception as e:
        return {"Status": "Error", "Message": str(e)}